from flask import Flask, send_file
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Color, Alignment, Border, Side, PatternFill

app = Flask(__name__)

def create_excel_file():
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "GenAI Readiness"

    # Sample Data for Sheet 1: General AI Readiness
    data1 = {
        "Category": ["Infrastructure and Data"] * 4,
        "Question": [
            "What existing data infrastructure supports AI projects?",
            "Is there a process for collecting, storing, and managing high-quality data?",
            "How does the current IT infrastructure support AI deployment?",
            "What data privacy and security measures are in place?"
        ]
    }

    df1 = pd.DataFrame(data1)
    for r in dataframe_to_rows(df1, index=False, header=True):
        ws1.append(r)

    # Sample Data for Sheet 2: Technical and Talent Readiness
    ws2 = wb.create_sheet(title="Technical and Talent Readiness")
    data2 = {
        "Category": ["Skills and Capabilities"] * 4,
        "Question": [
            "What is the current level of AI and machine learning expertise within the organization?",
            "Are there partnerships or collaborations with AI research institutions or companies?",
            "What training programs are available for staff to upgrade their AI skills?",
            "How is the organization staying updated with the latest advancements in AI?"
        ]
    }

    df2 = pd.DataFrame(data2)
    for r in dataframe_to_rows(df2, index=False, header=True):
        ws2.append(r)

    # Sample Data for Sheet 3: GenAI Specific Readiness by Area
    ws3 = wb.create_sheet(title="GenAI Specific Readiness")
    data3 = {
        "Category": ["Text", "Image", "Sound", "Voice", "Video", "3D Objects", "VR", "AR", "XR"] * 4,
        "Question": [
            "What are the potential applications of GenAI in this area?",
            "Do we have access to the necessary datasets to train GenAI models in this domain?",
            "What are the specific technical challenges in applying GenAI in this area?",
            "How will GenAI projects in this area align with the overall business strategy?"
        ] * 9
    }

    df3 = pd.DataFrame(data3)
    for r in dataframe_to_rows(df3, index=False, header=True):
        ws3.append(r)

    # Formatting for the first sheet as an example
    for cell in ws1["1:1"]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    filename = "Genesis_Project_Assessment.xlsx"
    wb.save(filename)
    print("Excel file created and saved successfully.")
    return filename

@app.route('/')
def download_excel():
    filename = create_excel_file()
    return send_file(filename, as_attachment=True)

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=8080)
